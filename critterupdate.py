import pandas
import openpyxl

def _or_in_(alist,astring):
	#function to determine if any words in a list (alist)
	#are present anywhere in a string (astring)
	
	for word in alist:
		if word.lower() in astring.lower():
			return True
	return False
def critterupdate():
        wbkbugs = 'insectlist.xlsx'
        wbkb = openpyxl.load_workbook(wbkbugs)
        insects_xl=pandas.read_excel(wbkbugs)
        insects=insects_xl.to_dict('records')
        bugnamelist=list(map(lambda x:(x['Name']),insects))
        sheetb=wbkb.worksheets[0]

        wbkfish = 'Fishlist.xlsx'
        wbkf = openpyxl.load_workbook(wbkfish)
        fish_xl=pandas.read_excel(wbkfish)
        fish=fish_xl.to_dict('records')
        fishnamelist=list(map(lambda x:(x['Name']),fish))
        sheetf=wbkf.worksheets[0]




                
        contsearch=True #Will continue to search until user declines to continue
        while contsearch:

                print ("What critter? ")
                whichcrit=input()
                bugnum=0
                fishnum=0
                if whichcrit.title() in bugnamelist:
                        bugnum=bugnamelist.index(whichcrit.title())+2

                elif whichcrit.title() in fishnamelist:
                        fishnum=fishnamelist.index(whichcrit.title())+2

                else:
                        print ("Is it a fish or a bug? ")
                        myinp=input()
                        if _or_in_(["bug","insect"],myinp):
                                
                                for bug in bugnamelist:
                                        if whichcrit.lower() in bug.lower():
                                                print ("Is this your bug? "+bug)
                                                check=input()
                                                if "y" in check.lower():
                                                        bugnum=bugnamelist.index(bug)+2
                                                        break
                        elif "fish" in myinp:
                                for fish in fishnamelist:
                                        if whichcrit.lower() in fish.lower():
                                                print ("Is this your fish? "+fish)
                                                check=input()
                                                if "y" in check.lower():
                                                        fishnum=fishnamelist.index(fish)+2
                                                        break
                alreadyasked=False
                if bugnum:
                        if sheetb.cell(row=bugnum,column=7).value=="Yes":
                                print ("Already in spreadsheet as donated.")
                        else:
                                sheetb.cell(row=bugnum,column=7).value="Yes"
                                if sheetb.cell(row=bugnum,column=7).value=="Yes":
                                        print ("Successfully added")
                elif fishnum:
                        if sheetf.cell(row=fishnum,column=8).value=="Yes":
                                print ("Already in spreadsheet as donated.")
                        else:
                                sheetf.cell(row=fishnum,column=8).value="Yes"
                                if sheetf.cell(row=fishnum,column=8).value=="Yes":
                                        print ("Successfully added")
                
                else:
                        print ("Sorry couldn't find that critter. Look for a different term? ")
                        searchcheck=input()
                        alreadyasked=True
                        if "y" not in searchcheck.lower():
                                contsearch=False
                if not alreadyasked:
                        
                        print ("Add another? ")
                        searchcheck=input()
                        if "y" not in searchcheck.lower():
                                contsearch=False

                                        
                        


        wbkb.save(wbkbugs)
        wbkb.close

        wbkf.save(wbkfish)
        wbkf.close

if __name__ == "__main__":
        critterupdate()
