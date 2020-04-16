import openpyxl
import pandas
import sys

def _or_in_(alist,astring):
	for word in alist:
		if word.lower() in astring.lower():
			return True
	return False
def endearly(astring):
	endlist=["nvm","cancel","stop","nevermind"]
	for word in endlist:
	    if word in astring.lower(): sys.exit()
	    
def fixacritter():
        myinp=input("Do you need to fix a bug entry or a fish entry? ")

        if _or_in_(["bug","insect"],myinp.lower()):
                wbkbugs = 'insectlist.xlsx'
                wbkb = openpyxl.load_workbook(wbkbugs)
                insects_xl=pandas.read_excel(wbkbugs)
                insects=insects_xl.to_dict('records')
                bugnamelist=list(map(lambda x:(x['Name']),insects))
                sheetb=wbkb.worksheets[0]
                whichbug=input("What bug? ")
                endearly(whichbug)
                keysl=list(insects[0].keys())[1:]
                if whichbug.title() in bugnamelist:
                        bugnum=bugnamelist.index(whichbug.title())+2
                else:
                        bsc=True
                        while bsc:
                                searchcheck=input("Sorry, couldn't find that bug. Try a search? ")
                                if "y" in searchcheck.lower():
                                        bugsearch=input("Search for what term? ")
                                        endearly(bugsearch)
                                        for buug in bugnamelist:
                                                if bugsearch.lower() in buug.lower():
                                                        buugcheck=input("Is this your bug? "+buug+" ")
                                                        if "y" in buugcheck.lower():
                                                                bugnum=bugnamelist.index(buug)+2
                                                                bsc=False
                                                                break
                                else: sys.exit()
                fixwhile=True
                while fixwhile:
                        fixer=input("Fix what part of the entry? Options: "+', '.join(keysl)+" ").capitalize()
                        endearly(fixer)
                        if fixer in keysl:
                                fixnum=keysl.index(fixer)+2
                                fixwhile=False
                newval=input("This cell currently reads: "+str(sheetb.cell(row=bugnum,column=fixnum).value)+". What should it read instead? ")
                endearly(newval)
                sheetb.cell(row=bugnum,column=fixnum).value=newval
                if sheetb.cell(row=bugnum,column=fixnum).value==newval:
                        print (sheetb.cell(row=bugnum,column=2).value+" "+fixer.lower()+" successfully changed to "+newval)

                wbkb.save(wbkbugs)
                wbkb.close
                
                
        if ("fish") in myinp.lower():
                wbkfish = 'Fishlist.xlsx'
                wbkf = openpyxl.load_workbook(wbkfish)
                fish_xl=pandas.read_excel(wbkfish)
                fish=fish_xl.to_dict('records')
                fishnamelist=list(map(lambda x:(x['Name']),fish))
                sheetf=wbkf.worksheets[0]
                whichfish=input("What fish? ")
                endearly(whichfish)
                keysl=list(fish[0].keys())[1:]
                if whichfish.title() in fishnamelist:
                        fishnum=fishnamelist.index(whichfish.title())+2
                else:
                        fsc=True
                        while fsc:
                                searchcheck=input("Sorry, couldn't find that fish. Try a search? ")
                                if "y" in searchcheck.lower():
                                        fishsearch=input("Search for what term? ")
                                        endearly(fishsearch)
                                        for fysh in fishnamelist:
                                                if fishsearch.lower() in fysh.lower():
                                                        fyshcheck=input("Is this your fish? "+fysh+" ")
                                                        if "y" in fyshcheck.lower():
                                                                fishnum=fishnamelist.index(fysh)+2
                                                                fsc=False
                                                                break
                                else: sys.exit()
                fixwhile=True
                while fixwhile:
                        fixer=input("Fix what part of the entry? Options: "+', '.join(keysl)+" ").capitalize()
                        endearly(fixer)
                        if fixer in keysl:
                            fixnum=keysl.index(fixer)+2
                            fixwhile=False
                                
                newval=input("This cell currently reads: "+str(sheetf.cell(row=fishnum,column=fixnum).value)+". What should it read instead? ")
                endearly(newval)
                sheetf.cell(row=fishnum,column=fixnum).value=newval
                if sheetf.cell(row=fishnum,column=fixnum).value==newval:
                        print (sheetf.cell(row=fishnum,column=2).value+" "+fixer.lower()+" successfully changed to "+newval)

                wbkf.save(wbkfish)
                wbkf.close

if __name__ == "__main__":
        fixacritter()
